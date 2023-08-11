import time
import pandas as pd

from openpyxl import load_workbook

russian_headings = {
    "services": 'Услуги',
    "hot_water": 'Вода гор. (куб.)',
    "cold_water": 'Вода хол. (куб.)',
    "electricity": 'Эл/э',
    "water_removal": 'Водоотведение (куб.)',
    "prepayment": 'Аванс',
    "internet": 'Интернет',
    "debt": 'Итого начислено',
    "payment": 'Оплата',
    "outcome": 'Итог',
    "previous": 'Предыдущее',
    "current": 'Текущее',
    "expenditure": 'Расход',
    "tariff": 'Тариф',
    "accrued_accor_t_t_tariff": 'Начислено по тарифу',
    "date": 'Дата',
    "info": 'Информация'
}

def return_formulas(shape):
    first_formulas = {
        "previous": [None, 0, 0, 0, f'=D3+D4', 1, 1, f'=SUM(F3:F8)', None, f'=B10-B9'],
        "expenditure": [None, f'=C3-B3', f'=C4-B4', f'=C5-B5', None, None, None, None, None, None],
        "tariff": [None, 0, 0, 0, 0, 0, 0, None, None, None],
        "accrued_accor_t_t_tariff": [None, f'=D3*E3', f'=D4*E4', f'=D5*E5', f'=B6*E6',
                                     f'=B7*E7', f'=B8*E8', None, None, None],
    }

    next_formulas = {
        "previous": [None, f"=C{shape - 7}", f"=C{shape - 6}", f"=C{shape - 5}",
                     f'=D{shape + 3}+D{shape + 4}', 1, 1, f'=SUM(F{shape + 3}:F{shape + 8})',
                     None, f'=B{shape + 10}-B{shape + 9}+B{shape + 1}'],
        "expenditure": [None, f'=C{shape + 3}-B{shape + 3}', f'=C{shape + 4}-B{shape + 4}',
                        f'=C{shape + 5}-B{shape + 5}', None, None, None, None, None, None],
        "tariff": [None, f"=E{shape - 7}", f"=E{shape - 6}", f"=E{shape - 5}", f"=E{shape - 4}",
                   f"=E{shape - 3}", f"=E{shape - 2}", None, None, None],
        "accrued_accor_t_t_tariff": [None, f'=D{shape + 3}*E{shape + 3}', f'=D{shape + 4}*E{shape + 4}',
                                     f'=D{shape + 5}*E{shape + 5}', f'=B{shape + 6}*E{shape + 6}',
                                     f'=B{shape + 7}*E{shape + 7}', f'=B{shape + 8}*E{shape + 8}',
                                     None, None, None]
    }

    if shape == 0:
        return first_formulas
    else:
        return next_formulas

def return_frame_with_heeadings(headings, shape):
    df = pd.DataFrame({headings["services"]: [time.strftime("%Y-%m"),
                                               headings["hot_water"],
                                               headings["cold_water"],
                                               headings["electricity"],
                                               headings["water_removal"],
                                               headings["prepayment"],
                                               headings["internet"],
                                               headings["debt"],
                                               headings["payment"],
                                               headings["outcome"]],
                       headings["previous"]: return_formulas(shape)["previous"],
                       headings["current"]:   [None, 0, 0, 0, None, None, None, None, None, None],
                       headings["expenditure"]: return_formulas(shape)["expenditure"],
                       headings["tariff"]:   return_formulas(shape)["tariff"],
                       headings["accrued_accor_t_t_tariff"]: return_formulas(shape)["accrued_accor_t_t_tariff"],
                       headings["date"]: [None, None, None, None, None,
                                None, None, None, None, None],
                       headings["info"]: [None, None, None, None, None,
                                None, None, None, None, None]
                       })
    return df

def return_df_from_excel(xl, sheet):
    #xl = load_workbook(path)
    df = pd.DataFrame(xl[sheet].values)
    if df.shape[0] != 0:
        df.columns = df.iloc[0]
        df = df.drop(index=0)
    return df

def insert_new(xl, writer):
    for sheet in xl.sheetnames:
        df = return_df_from_excel(xl, sheet)
        if df.shape[0] == 0:
            return_frame_with_heeadings(russian_headings, 0).to_excel(writer, index=False, sheet_name=sheet)
        else:
            if df.iat[df.shape[0] - 10, 0] != time.strftime("%Y-%m"):
                df = pd.concat([df, return_frame_with_heeadings(russian_headings, df.shape[0])], axis=0)
                df.to_excel(writer, index=False, sheet_name=sheet)

            else:
                df.to_excel(writer, index=False, sheet_name=sheet)
    writer.close()

def edit_last(xl):
    for sheet in xl.sheetnames:
        df = return_df_from_excel(xl, sheet)

        ws = xl[sheet]
        ws.merge_cells(f'A{df.shape[0] - 8}:F{df.shape[0] - 8}')
        ws.merge_cells(f'B{df.shape[0] - 4}:D{df.shape[0] - 4}')
        ws.merge_cells(f'B{df.shape[0] - 3}:D{df.shape[0] - 3}')
        ws.merge_cells(f'B{df.shape[0] - 2}:D{df.shape[0] - 2}')
        ws.merge_cells(f'B{df.shape[0] - 1}:F{df.shape[0] - 1}')
        ws.merge_cells(f'B{df.shape[0]}:F{df.shape[0]}')
        ws.merge_cells(f'B{df.shape[0] + 1}:F{df.shape[0] + 1}')